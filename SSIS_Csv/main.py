from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow,QTableWidgetItem, QTableWidget, QWidget, QMessageBox
from PyQt5.QtCore import Qt, QPoint, QRect, QPropertyAnimation, QEvent, QRegExp
from PyQt5.QtGui import QRegExpValidator

import sys
import openpyxl

from SSIS import Ui_MainWindow

class SSIS_Window(QMainWindow):
    def __init__(self):
        super().__init__()
    #Window Setup + Frameless
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setWindowIcon(QtGui.QIcon('Icons/Window Icon.ico'))
        self.setWindowTitle("Student Information System")
    #Initializer
        self.initializeButtons()
        self.load_data()
        self.load_everything_else()
        self.ui.Editbtn.hide()  
        self.ui.Deletebtn.hide() 
        self.ui.Deselectbtn.hide()
        self.set_boxtext()
    #Animation variables
        self.Sidemenu_counter = 0
        self.Sidemenu_parameter = QRect(181,30,899,693)
        self.Searchbtn_counter = 0
        self.search_x = self.ui.Search_Interface.geometry().x()
        self.search_y = self.ui.Search_Interface.geometry().y()
        self.Searchbtn_parameter = QRect(self.search_x,self.search_y,899,693)
        self.Addbtn_counter = 0
        self.add_x = self.ui.Add_Interface.geometry().x()
        self.add_y = self.ui.Add_Interface.geometry().y()
        self.Addbtn_parameter = QRect(self.add_x,self.add_y,899,693)
        self.Editbtn_counter = 0
        self.edit_x = self.ui.Edit_Interface.geometry().x()
        self.edit_y = self.ui.Edit_Interface.geometry().y()
        self.Editbtn_parameter = QRect(self.edit_x,self.edit_y,899,693)
        self.Main_parameter = QRect(0, 28, 1101, 701)
    #Window Moving variables
        self.old_pos = self.pos()
        self.title_parameter = QRect(0,0,1000,31)
        self.mouse_pressed = False
        self.mouse_tracking = False
    
    def Refresh_Table(self):
        self.ui.Studenttable.setRowCount(0)
        self.load_data()

    def Add_Student_Info(self):
        self.course_mapping = {
            "Bachelor of Science in Chemical Engineering": "BSChE",
            "Bachelor of Science in Environmental Engineering": "BSEnE",
            "Bachelor of Science in Civil Engineering": "BSCE",
            "Bachelor of Science in Computer Engineering": "BSCpE",
            "Bachelor of Science in Electrical Engineering": "BSEE",
            "Bachelor of Science in Electronics and Communication Engineering": "BSECE",
            "Bachelor of Science in Industrial Automation and Mechatronics": "BSIAM",
            "Bachelor of Science in Ceramics Engineering": "BSCerE",
            "Bachelor of Science in Mechanical Engineering": "BSME",
            "Bachelor of Engineering Technology Major in Chemical Engineering and Technology": "BET-ChE",
            "Bachelor of Engineering Technology Major in Civil Engineering Technology": "BET-CET",
            "Bachelor of Engineering Technology Major in Electrical Engineering Technology": "BET-EET",
            "Bachelor of Engineering Technology Major in Electronics Engineering Technology": "BET-EST",
            "Bachelor of Engineering Technology Major in Metallurgical and Materials Engineering Technology": "BET-MMT",
            "Bachelor of Engineering Technology Major in Mechanical Engineering Technology": "BET-MET",
            "Bachelor of Science in Biology": "BS Biology",
            "Bachelor of Science in Chemistry": "BS Chemistry",
            "Bachelor of Science in Mathematics": "BS Mathematics",
            "Bachelor of Science in Statistics": "BS Statistics",
            "Bachelor of Science in Physics": "BS Physics",
            "Bachelor of Science in Computer Science": "BSCS",
            "Bachelor of Science in Information Technology": "BSIT",
            "Bachelor of Science in Information System": "BSIS",
            "Bachelor of Science in Computer Application": "BSCA",
            "Bachelor of Elementary Education in Science and Mathematics": "BEED Science and Mathematics",
            "Bachelor of Elementary Education in Language Education": "BEED Language Education",
            "Bachelor of Secondary Education in Chemistry": "BSEd Chemistry",
            "Bachelor of Secondary Education in Physics": "BSEd Physics",
            "Bachelor of Secondary Education in Mathematics": "BSEd Mathematics",
            "Bachelor of Secondary Education in Biology": "BSEd Biology",
            "Bachelor of Secondary Education in Filipino": "BSEd Filipino",
            "Bachelor of Technology and Livelihood in Home Economics": "BTLEd Home Economics",
            "Bachelor of Technology and Livelihood in Industrial Arts": "BTLEd Industrial Arts",
            "Bachelor of Technical-Vocational Teacher Education in Drafting Technology": "BTVTEd Drafting",
            "Bachelor of Arts in English Language Studies": "BAELS",
            "Bachelor of Arts in Language and Culture Studies": "BALCS",
            "Bachelor of Arts in Filipino": "BA Filipino",
            "Bachelor of Arts in History": "BA History",
            "Bachelor of Arts in Panitikan": "BA Panitikan",
            "Bachelor of Arts in Political Science": "BA Political Science",
            "Bachelor of Arts in Sociology": "BA Sociology",
            "Bachelor of Arts in Psychology": "BA Psychology",
            "Bachelor of Science in Psychology": "BS Psychology",
            "Bachelor of Science in Philosophy": "BS Philosophy",
            "Bachelor of Science in Accountancy": "BSA",
            "Bachelor of Science in Economics": "BS Economics",
            "Bachelor of Science in Business Administration Major in Business Economics": "BSBA-BE",
            "Bachelor of Science in Entrepreneurship": "BSENTREP",
            "Bachelor of Science in Hotel Management": "BSHM",
            "Bachelor of Science in Nursing": "BSN"
        }
        Id_validator = QRegExpValidator(QRegExp(r"^\d{4}-\d{4}$"))
        self.ui.IDNUM_Text.setValidator(Id_validator)

        id_number = self.ui.IDNUM_Text.text().strip()
        first_name = self.ui.FIRSTNAME_Text.text().strip().title()
        last_name = self.ui.LASTNAME_Text.text().strip().title()
        gender = self.ui.GENDER_Box.currentText().strip()
        year_level = self.ui.YEARLEVEL_Box.currentText().strip()
        course = self.ui.COURSE_Box.currentText().strip()
        course = self.course_mapping.get(course, course)
        
        if not id_number and not first_name and not last_name and not gender and not year_level and not course:
            self.show_error_message("All fields must be filled!")
            return 
        if not id_number:
            self.show_error_message("Please put an ID Number!")
            return 
        if not first_name:
            self.show_error_message("Please put first name!")
            return 
        if not last_name:
            self.show_error_message("PLease put last name!")
            return 
        if len(id_number) != 9 or not id_number.startswith("2") or "-" not in id_number:
            self.show_error_message("ID Number is not in the correct format!\n Correct format: 2NNN-NNNN")
            return
        if self.ID_Duplicate_Checker(id_number):
            self.show_error_message("This ID Number already exists!")
            return
        if self.ui.GENDER_Box.currentIndex() == 0:
            self.show_error_message("Please select a gender!")
            return
        if self.ui.COURSE_Box.currentIndex() == 0:
            self.show_error_message("Please select a course!")
            return
        if self.ui.YEARLEVEL_Box.currentIndex() == 0:
            self.show_error_message("Please select a year level!")
            return

        # Add new row to table if ID is unique
        row_position = self.ui.Studenttable.rowCount()
        self.ui.Studenttable.insertRow(row_position)
        self.ui.Studenttable.setItem(row_position, 0, QtWidgets.QTableWidgetItem(id_number))
        self.ui.Studenttable.setItem(row_position, 1, QtWidgets.QTableWidgetItem(first_name))
        self.ui.Studenttable.setItem(row_position, 2, QtWidgets.QTableWidgetItem(last_name))
        self.ui.Studenttable.setItem(row_position, 3, QtWidgets.QTableWidgetItem(year_level))
        self.ui.Studenttable.setItem(row_position, 4, QtWidgets.QTableWidgetItem(gender))
        self.ui.Studenttable.setItem(row_position, 5, QtWidgets.QTableWidgetItem(course))
        self.show_success_message("Student Information successfully added!")
        self.ui.IDNUM_Text.clear()
        self.ui.FIRSTNAME_Text.clear()
        self.ui.LASTNAME_Text.clear()
        self.ui.GENDER_Box.setCurrentIndex(0)
        self.ui.YEARLEVEL_Box.setCurrentIndex(0)
        self.ui.COURSE_Box.setCurrentIndex(0)
        self.save_data()
        self.Refresh_Table()

    def Edit_Student_Info(self):
        self.course_mapping = {
            "Bachelor of Science in Chemical Engineering": "BSChE",
            "Bachelor of Science in Environmental Engineering": "BSEnE",
            "Bachelor of Science in Civil Engineering": "BSCE",
            "Bachelor of Science in Computer Engineering": "BSCpE",
            "Bachelor of Science in Electrical Engineering": "BSEE",
            "Bachelor of Science in Electronics and Communication Engineering": "BSECE",
            "Bachelor of Science in Industrial Automation and Mechatronics": "BSIAM",
            "Bachelor of Science in Ceramics Engineering": "BSCerE",
            "Bachelor of Science in Mechanical Engineering": "BSME",
            "Bachelor of Engineering Technology Major in Chemical Engineering and Technology": "BET-ChE",
            "Bachelor of Engineering Technology Major in Civil Engineering Technology": "BET-CET",
            "Bachelor of Engineering Technology Major in Electrical Engineering Technology": "BET-EET",
            "Bachelor of Engineering Technology Major in Electronics Engineering Technology": "BET-EST",
            "Bachelor of Engineering Technology Major in Metallurgical and Materials Engineering Technology": "BET-MMT",
            "Bachelor of Engineering Technology Major in Mechanical Engineering Technology": "BET-MET",
            "Bachelor of Science in Biology": "BS Biology",
            "Bachelor of Science in Chemistry": "BS Chemistry",
            "Bachelor of Science in Mathematics": "BS Mathematics",
            "Bachelor of Science in Statistics": "BS Statistics",
            "Bachelor of Science in Physics": "BS Physics",
            "Bachelor of Science in Computer Science": "BSCS",
            "Bachelor of Science in Information Technology": "BSIT",
            "Bachelor of Science in Information System": "BSIS",
            "Bachelor of Science in Computer Application": "BSCA",
            "Bachelor of Elementary Education in Science and Mathematics": "BEED Science and Mathematics",
            "Bachelor of Elementary Education in Language Education": "BEED Language Education",
            "Bachelor of Secondary Education in Chemistry": "BSEd Chemistry",
            "Bachelor of Secondary Education in Physics": "BSEd Physics",
            "Bachelor of Secondary Education in Mathematics": "BSEd Mathematics",
            "Bachelor of Secondary Education in Biology": "BSEd Biology",
            "Bachelor of Secondary Education in Filipino": "BSEd Filipino",
            "Bachelor of Technology and Livelihood in Home Economics": "BTLEd Home Economics",
            "Bachelor of Technology and Livelihood in Industrial Arts": "BTLEd Industrial Arts",
            "Bachelor of Technical-Vocational Teacher Education in Drafting Technology": "BTVTEd Drafting",
            "Bachelor of Arts in English Language Studies": "BAELS",
            "Bachelor of Arts in Language and Culture Studies": "BALCS",
            "Bachelor of Arts in Filipino": "BA Filipino",
            "Bachelor of Arts in History": "BA History",
            "Bachelor of Arts in Panitikan": "BA Panitikan",
            "Bachelor of Arts in Political Science": "BA Political Science",
            "Bachelor of Arts in Sociology": "BA Sociology",
            "Bachelor of Arts in Psychology": "BA Psychology",
            "Bachelor of Science in Psychology": "BS Psychology",
            "Bachelor of Science in Philosophy": "BS Philosophy",
            "Bachelor of Science in Accountancy": "BSA",
            "Bachelor of Science in Economics": "BS Economics",
            "Bachelor of Science in Business Administration Major in Business Economics": "BSBA-BE",
            "Bachelor of Science in Entrepreneurship": "BSENTREP",
            "Bachelor of Science in Hotel Management": "BSHM",
            "Bachelor of Science in Nursing": "BSN"
        }
        Id_validator = QRegExpValidator(QRegExp(r"^\d{4}-\d{4}$"))
        selected_row = self.ui.Studenttable.currentRow()
        self.ui.IDNUM2_Text.setValidator(Id_validator)
        table_id = self.ui.Studenttable.item(selected_row, 0).text().strip()
        input_id = self.ui.IDNUM_Text.text().strip()

        id_number = self.ui.IDNUM2_Text.text().strip()
        first_name = self.ui.FIRSTNAME2_Text.text().strip().title()
        last_name = self.ui.LASTNAME2_Text.text().strip().title()
        gender = self.ui.GENDER2_Box.currentText().strip()
        year_level = self.ui.YEARLEVEL2_Box.currentText().strip()
        course = self.ui.COURSE2_Box.currentText().strip()
        old_gender = self.ui.GENDER3_Box.text().strip()
        old_year_level = self.ui.YEARLEVEL3_Box.text().strip()
        old_course = self.ui.COURSE3_Box.text().strip()
        course = self.course_mapping.get(course, course)
        
        if not id_number and not first_name and not last_name and not gender and not year_level and not course:
            self.show_error_message("All fields must be filled!")
            return 
        if not id_number:
            self.show_error_message("Please put an ID Number!")
            return 
        if not first_name:
            self.show_error_message("Please put first name!")
            return 
        if not last_name:
            self.show_error_message("PLease put last name!")
            return 
        if len(id_number) != 9 or not id_number.startswith("2") or "-" not in id_number:
            self.show_error_message("ID Number is not in the correct format!\n Correct format: 2NNN-NNNN")
            return
        if table_id == input_id:
            pass
            if self.ID_Duplicate_Checker(id_number):
                self.show_error_message("This ID Number already exists!")
                return
        
        

        
        self.ui.Studenttable.setItem(selected_row, 0, QtWidgets.QTableWidgetItem(id_number))
        self.ui.Studenttable.setItem(selected_row, 1, QtWidgets.QTableWidgetItem(first_name))
        self.ui.Studenttable.setItem(selected_row, 2, QtWidgets.QTableWidgetItem(last_name))
        if self.ui.GENDER2_Box.currentIndex() == 0:
            self.ui.Studenttable.setItem(selected_row, 4, QtWidgets.QTableWidgetItem(old_gender))
        else:
            self.ui.Studenttable.setItem(selected_row, 4, QtWidgets.QTableWidgetItem(gender))
        if self.ui.YEARLEVEL2_Box.currentIndex() == 0:
            self.ui.Studenttable.setItem(selected_row, 3, QtWidgets.QTableWidgetItem(old_year_level))
        else:
            self.ui.Studenttable.setItem(selected_row, 3, QtWidgets.QTableWidgetItem(year_level))
        if self.ui.COURSE2_Box.currentIndex() == 0:
            self.ui.Studenttable.setItem(selected_row, 5, QtWidgets.QTableWidgetItem(old_course))
        else:
            self.ui.Studenttable.setItem(selected_row, 5, QtWidgets.QTableWidgetItem(course))

        self.show_success_message("Student Information successfully added!")
        self.deselect_row()
        self.ui.IDNUM2_Text.clear()
        self.ui.FIRSTNAME2_Text.clear()
        self.ui.LASTNAME2_Text.clear()
        self.ui.GENDER2_Box.setCurrentIndex(0)
        self.ui.YEARLEVEL2_Box.setCurrentIndex(0)
        self.ui.COURSE2_Box.setCurrentIndex(0)
        self.save_data()
        self.Refresh_Table()

    def ID_Duplicate_Checker(self, id_number):
        for row in range(self.ui.Studenttable.rowCount()):
            item = self.ui.Studenttable.item(row, 0)
            if item and item.text() == id_number:
                return True
        return False  

    #for sorting
    #def Table_hide(self):

    #def Sort_by(self):

    #def Edit_Info(self):
    

    def Search_table(self):
        search = self.ui.SearchLineEdit.text().strip().lower()

        if not search:
            for row in range(self.ui.Studenttable.rowCount()):
                self.ui.Studenttable.setRowHidden(row, False)
            return

        for row in range(self.ui.Studenttable.rowCount()):
            self.row_match = False

            for col in range(self.ui.Studenttable.columnCount()):
                item = self.ui.Studenttable.item(row,col)

                if item and search in item.text().strip().lower():
                    self.row_match = True
                    break

            self.ui.Studenttable.setRowHidden(row, not self.row_match)
        self.ui.Studenttable.sortItems(1, QtCore.Qt.AscendingOrder)
        self.ui.SearchLineEdit.clear()

    #Save Data to Excel Function
    def save_data(self):
        path = "Student Information.xlsx"
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Clear old data (except headers)
        ws.delete_rows(2, ws.max_row)  # Remove everything from row 2 onwards
        # Get table row and column counts
        row_count = self.ui.Studenttable.rowCount()
        col_count = self.ui.Studenttable.columnCount()
        # Write new data from table to Excel
        for row in range(row_count):
            row_data = []
            for col in range(col_count):
                item = self.ui.Studenttable.item(row, col)
                row_data.append(item.text() if item else "")  # Handle empty cells
            ws.append(row_data)  # Append new row to Excel sheet
        wb.save(path)
        wb.close()
        self.show_success_message("Data saved successfully to Excel!")
        self.Refresh_Table()

    #Load Excel Data Function
    def load_data (self):
        path = "Student Information.xlsx"
        wb_student =  openpyxl.load_workbook(path)
        wb_student.active = wb_student['Student']
        student = wb_student.active
        
        self.ui.Studenttable.setRowCount(student.max_row)
        self.ui.Studenttable.setColumnCount(student.max_column)
        self.ui.Studenttable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.ui.Studenttable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.ui.Studenttable.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        list_values = list(student.values)
        non_empty_rows = [row for row in list_values if any(row)]
        self.ui.Studenttable.setHorizontalHeaderLabels(list_values[0])
        row_index = 0      
        for value_tuple in non_empty_rows[1:]:
            column_index = 0
            for value in value_tuple:
                value = str(value) if value is not None else " "
                self.ui.Studenttable.setItem(row_index, column_index, QTableWidgetItem(str(value)))
                column_index += 1
            row_index += 1
        self.remove_empty_rows()

    def remove_empty_rows(self):
        row_count = self.ui.Studenttable.rowCount()
        col_count = self.ui.Studenttable.columnCount()

        for row in range(row_count -1, -1, -1):  
            is_empty = True

            for col in range(col_count):
                item = self.ui.Studenttable.item(row, col)
                if item and item.text().strip():  
                    is_empty = False
                    break  
        
            if is_empty:
                self.ui.Studenttable.removeRow(row)  

    def Delete_row(self):
        selected_row = self.ui.Studenttable.currentRow()  # Get selected row index

        self.confirmation = self.show_question_message("Are you sure?")

        if self.confirmation == True:
            self.ui.Studenttable.removeRow(selected_row)  # Remove row and shift up
            self.show_success_message("Row deleted successfully!")
            self.save_data()  # Save changes to Excel

    #Initialize All Button
    def initializeButtons(self):
        #Menu Animation
        self.ui.Menubtn.clicked.connect(lambda: (self.Sidemenu_Animation(),self.deselect_row(), self.Editbtn_Animation_close()))
        self.ui.Menubtn_menu.clicked.connect(lambda: (self.Sidemenu_Animation(),self.deselect_row(), self.Editbtn_Animation_close()))
        #Search and Add Animation
        self.ui.Searchbtn.clicked.connect(lambda: (self.Searchbtn_Animation(),self.deselect_row()))
        self.ui.Addbtn.clicked.connect(lambda: (self.Addbtn_Animation(),self.deselect_row()))
        self.ui.Editbtn.clicked.connect(lambda:self.Editbtn_Animation())
        #Minimize and Quit
        self.ui.Minimizebtn.clicked.connect(self.showMinimized)
        self.ui.Quitbtn.clicked.connect(self.close)
        #Stacked Widget
        self.ui.Addbtn.clicked.connect(lambda: self.ui.Table.setCurrentWidget(self.ui.Student))
        self.ui.Searchbtn.clicked.connect(lambda: self.ui.Table.setCurrentWidget(self.ui.Student))
        self.ui.Studentbtn.clicked.connect(lambda: self.ui.Table.setCurrentWidget(self.ui.Student))
        self.ui.Programbtn.clicked.connect(lambda: self.ui.Table.setCurrentWidget(self.ui.Program))
        self.ui.Collegebtn.clicked.connect(lambda: self.ui.Table.setCurrentWidget(self.ui.College))    
        #Toggle Edit and Delete Button  
        self.ui.Studenttable.itemSelectionChanged.connect(lambda: self.toggle_delete_and_add_button())
        self.ui.Deselectbtn.clicked.connect(lambda: self.deselect_row())
        self.ui.Studenttable.selectionModel().selectionChanged.connect(lambda: (self.on_selectionChanged, self.Searchbtn_Animation_close(), self.Addbtn_Animation_close(), self.Sidemenu_Animation_close(), self.load_selected_row_data()))
        #Add Info
        self.ui.Addinfobtn.clicked.connect(lambda: self.Add_Student_Info())
        #Edit Info
        self.ui.Editinfobtn.clicked.connect(lambda: self.Edit_Student_Info())
        #Refresh Button
        self.ui.Refreshbtn.clicked.connect(lambda: (self.deselect_row(),self.Refresh_Table()))
        #Table Sort
        #Table Search
        self.ui.Searchlinebtn.clicked.connect(lambda: self.Search_table())
        #Delete
        self.ui.Deletebtn.clicked.connect(lambda: self.Delete_row())

    def load_selected_row_data(self):
        selected_row = self.ui.Studenttable.currentRow()

        self.ui.IDNUM2_Text.setText(self.ui.Studenttable.item(selected_row, 0).text())  
        self.ui.FIRSTNAME2_Text.setText(self.ui.Studenttable.item(selected_row, 1).text())  
        self.ui.LASTNAME2_Text.setText(self.ui.Studenttable.item(selected_row, 2).text())  
        self.ui.GENDER3_Box.setText(self.ui.Studenttable.item(selected_row, 3).text())  
        self.ui.YEARLEVEL3_Box.setText(self.ui.Studenttable.item(selected_row, 4).text())  
        self.ui.COURSE3_Box.setText(self.ui.Studenttable.item(selected_row, 5).text())
    
    def on_selectionChanged(self, selected, deselected):
        for ix in selected.indexes():
            print('', format(ix.row()), ix.column())

        for ix in deselected.indexes():
            print('', format(ix.row()), ix.column())

    def toggle_delete_and_add_button(self):
        selected_rows = self.ui.Studenttable.selectionModel().selectedRows()
        if selected_rows:
            self.ui.Editbtn.show()
            self.ui.Deletebtn.show()
            self.ui.Deselectbtn.show()
        else:
            self.ui.Editbtn.hide()
            self.ui.Deletebtn.hide()
            self.ui.Deselectbtn.hide()

    def deselect_row(self):
        self.ui.Studenttable.clearSelection()
        self.ui.Collegetable.clearSelection()
        self.ui.Programtable.clearSelection()

    def show_error_message(self, message):
        self.msg = QtWidgets.QMessageBox()
        self.msg.setIcon(QtWidgets.QMessageBox.Critical) 
        self.msg.setWindowTitle("Error")  
        self.msg.setText(message)  
        self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)  
        self.msg.setStyleSheet("QLabel{color:red; font-size:14px; bold}")
        self.msg.exec_()  

    def show_question_message(self, message):
        self.msg = QtWidgets.QMessageBox()
        self.msg.setIcon(QtWidgets.QMessageBox.Question)
        self.msg.setWindowTitle("Confirmation") 
        self.msg.setText(message)  
        self.msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No) 
        self.msg.setStyleSheet("QLabel{color:white; font-size:14px; bold}")
        response = self.msg.exec_()

        if response == QtWidgets.QMessageBox.Yes:
            return True  
        else:
            return False  

    def show_success_message(self, message):
        self.msg = QtWidgets.QMessageBox()
        self.msg.setIcon(QtWidgets.QMessageBox.Information)
        self.msg.setWindowTitle("Success")  
        self.msg.setText(message)  
        self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)  
        self.msg.setStyleSheet("QLabel{color:white; font-size:14px; bold}")
        self.msg.exec_()  

    
    #Window Moving
    def mousePressEvent(self, event):
        self.old_pos = event.globalPos()
        self.posmouse = event.pos()
        self.mouse_pressed = True
        if self.title_parameter.contains(self.posmouse):
            self.mouse_tracking = True
        if self.Searchbtn_counter == 1 and self.Sidemenu_counter == 0 and self.Searchbtn_parameter.contains(self.posmouse):
            self.Searchbtn_Animation()
            self.Searchbtn_Animation_close()
        if self.Addbtn_counter == 1 and self.Sidemenu_counter == 0  and self.Addbtn_parameter.contains(self.posmouse):
            self.Addbtn_Animation()
            self.Addbtn_Animation_close()
        if self.Editbtn_counter == 1 and self.Sidemenu_counter == 0  and self.Editbtn_parameter.contains(self.posmouse):
            self.Editbtn_Animation()
            self.Editbtn_Animation_close()
        if self.Sidemenu_counter == 1 and self.Sidemenu_parameter.contains(self.posmouse):
            self.Sidemenu_Animation()
            self.Sidemenu_Animation_close()
        if self.Main_parameter.contains(self.posmouse):
            self.deselect_row()
        
    #Mouse Tracker
    def mouseMoveEvent(self, event): 
        if self.mouse_pressed and self.mouse_tracking:
            delta = QPoint(event.globalPos() - self.old_pos)
            self.move(self.x() + delta.x(), self.y() + delta.y()) 
            self.old_pos = event.globalPos()
    def mouseReleaseEvent(self, event):
        self.mouse_pressed = False
        self.mouse_tracking = False

    #Animation variables
    def Addbtn_Animation(self):
        if self.Addbtn_counter == 0:
            self.AddCurr1 = self.ui.Add_Interface.geometry()
            self.Add_Show = QtCore.QPropertyAnimation(self.ui.Add_Interface, b"pos")
            self.Add_Show.setDuration(200)
            self.Add_Show.setEndValue(QPoint(self.AddCurr1.x(),0))
            self.Add_Show.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Add_Show.start()
            self.Addbtn_counter = 1
    def Addbtn_Animation_close(self):
        if self.Addbtn_counter == 1:
            self.AddCurr2 = self.ui.Add_Interface.geometry()
            self.Add_Hide = QtCore.QPropertyAnimation(self.ui.Add_Interface, b"pos")
            self.Add_Hide.setDuration(500)
            self.Add_Hide.setEndValue(QPoint(self.AddCurr2.x(),-200))
            self.Add_Hide.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Add_Hide.start()
            self.Addbtn_counter = 0
    def Editbtn_Animation(self):
        if self.Editbtn_counter == 0:
            self.EditCurr1 = self.ui.Edit_Interface.geometry()
            self.Edit_Show = QtCore.QPropertyAnimation(self.ui.Edit_Interface, b"pos")
            self.Edit_Show.setDuration(200)
            self.Edit_Show.setEndValue(QPoint(self.EditCurr1.x(),0))
            self.Edit_Show.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Edit_Show.start()
            self.Editbtn_counter = 1
    def Editbtn_Animation_close(self):
        if self.Editbtn_counter == 1:
            self.EditCurr2 = self.ui.Edit_Interface.geometry()
            self.Edit_Hide = QtCore.QPropertyAnimation(self.ui.Edit_Interface, b"pos")
            self.Edit_Hide.setDuration(500)
            self.Edit_Hide.setEndValue(QPoint(self.EditCurr2.x(),-200))
            self.Edit_Hide.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Edit_Hide.start()
            self.Editbtn_counter = 0
    def Searchbtn_Animation(self):
        if self.Searchbtn_counter == 0:
            self.SearchCurr1 = self.ui.Search_Interface.geometry()
            self.Search_Show = QtCore.QPropertyAnimation(self.ui.Search_Interface, b"pos")
            self.Search_Show.setDuration(200)
            self.Search_Show.setEndValue(QPoint(self.SearchCurr1.x(),0))
            self.Search_Show.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Search_Show.start()
            self.Searchbtn_counter = 1
    def Searchbtn_Animation_close(self):
        if self.Searchbtn_counter == 1:
            self.SearchCurr2 = self.ui.Search_Interface.geometry()
            self.Search_Hide = QtCore.QPropertyAnimation(self.ui.Search_Interface, b"pos")
            self.Search_Hide.setDuration(500)
            self.Search_Hide.setEndValue(QPoint(self.SearchCurr2.x(),-61))
            self.Search_Hide.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.Search_Hide.start()
            self.Searchbtn_counter = 0

    def Sidemenu_Animation(self):
        if self.Sidemenu_counter == 0:
            self.AddCurr3 = self.ui.Add_Interface.geometry()
            self.SearchCurr3 = self.ui.Search_Interface.geometry()
            self.EditCurr3 = self.ui.Edit_Interface.geometry()
            self.animation_open = QtCore.QPropertyAnimation(self.ui.Sidemenu, b"pos")
            self.animation_open.setDuration(500)
            self.animation_open.setEndValue(QPoint(0,30))
            self.animation_open.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_open.start()
            self.animation_open2 = QtCore.QPropertyAnimation(self.ui.Search_Add, b"pos")
            self.animation_open2.setDuration(500)
            self.animation_open2.setEndValue(QPoint(195,0))
            self.animation_open2.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_open2.start()
            self.animation_open3 = QtCore.QPropertyAnimation(self.ui.Add_Interface, b"pos")
            self.animation_open3.setDuration(500)
            self.animation_open3.setEndValue(QPoint(195,self.AddCurr3.y()))
            self.animation_open3.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_open3.start()
            self.animation_open4 = QtCore.QPropertyAnimation(self.ui.Search_Interface, b"pos")
            self.animation_open4.setDuration(500)
            self.animation_open4.setEndValue(QPoint(195,self.SearchCurr3.y()))
            self.animation_open4.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_open4.start()
            self.animation_open5 = QtCore.QPropertyAnimation(self.ui.Edit_Interface, b"pos")
            self.animation_open5.setDuration(500)
            self.animation_open5.setEndValue(QPoint(195,self.EditCurr3.y()))
            self.animation_open5.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_open5.start()

            self.Sidemenu_counter = 1
    def Sidemenu_Animation_close(self):
        if self.Sidemenu_counter == 1:
            self.AddCurr4 = self.ui.Add_Interface.geometry()
            self.SearchCurr4 = self.ui.Search_Interface.geometry()
            self.EditCurr4 = self.ui.Edit_Interface.geometry()
            self.animation_close = QtCore.QPropertyAnimation(self.ui.Sidemenu, b"pos")
            self.animation_close.setDuration(500)
            self.animation_close.setEndValue(QPoint(-181,30))
            self.animation_close.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_close.start()
            self.animation_close2 = QtCore.QPropertyAnimation(self.ui.Search_Add, b"pos")
            self.animation_close2.setDuration(500)
            self.animation_close2.setEndValue(QPoint(75,0))
            self.animation_close2.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_close2.start()
            self.animation_close3 = QtCore.QPropertyAnimation(self.ui.Add_Interface, b"pos")
            self.animation_close3.setDuration(500)
            self.animation_close3.setEndValue(QPoint(75,self.AddCurr4.y()))
            self.animation_close3.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_close3.start()
            self.animation_close4 = QtCore.QPropertyAnimation(self.ui.Search_Interface, b"pos")
            self.animation_close4.setDuration(500)
            self.animation_close4.setEndValue(QPoint(75,self.SearchCurr4.y()))
            self.animation_close4.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_close4.start()
            self.animation_close5 = QtCore.QPropertyAnimation(self.ui.Edit_Interface, b"pos")
            self.animation_close5.setDuration(500)
            self.animation_close5.setEndValue(QPoint(75,self.EditCurr4.y()))
            self.animation_close5.setEasingCurve(QtCore.QEasingCurve.InOutCubic)
            self.animation_close5.start()
            self.Sidemenu_counter = 0
            

    def load_everything_else(self):
        wb_program =  openpyxl.load_workbook('Student Information.xlsx')
        wb_program.active = wb_program['Program']
        program = wb_program.active
        self.ui.Programtable.setRowCount(program.max_row)
        self.ui.Programtable.setColumnCount(program.max_column)
        self.ui.Programtable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.ui.Programtable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.ui.Programtable.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        list_values1 = list(program.values)
        self.ui.Programtable.setHorizontalHeaderLabels(list_values1[0])
        row_index1 = 0      
        for value_tuple1 in list_values1[1:]:
            column_index1 = 0
            for value in value_tuple1:
                value = str(value) if value is not None else " "
                self.ui.Programtable.setItem(row_index1, column_index1, QTableWidgetItem(str(value)))
                column_index1 += 1
            row_index1 += 1
        wb_college =  openpyxl.load_workbook('Student Information.xlsx')
        wb_college.active = wb_college['College']
        college = wb_college.active
        self.ui.Collegetable.setRowCount(college.max_row)
        self.ui.Collegetable.setColumnCount(college.max_column)
        self.ui.Collegetable.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.ui.Collegetable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.ui.Collegetable.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        list_values2 = list(college.values)
        self.ui.Collegetable.setHorizontalHeaderLabels(list_values2[0])
        row_index2 = 0      
        for value_tuple2 in list_values2[1:]:
            column_index2 = 0
            for value in value_tuple2:
                value = str(value) if value is not None else " "
                self.ui.Collegetable.setItem(row_index2, column_index2, QTableWidgetItem(str(value)))
                column_index2 += 1
            row_index2 += 1

    

    def set_boxtext(self):
        courses = ["Bachelor of Science in Chemical Engineering",
    "Bachelor of Science in Environmental Engineering", 
    "Bachelor of Science in Civil Engineering",
    "Bachelor of Science in Computer Engineering",
    "Bachelor of Science in Electrical Engineering",
    "Bachelor of Science in Electronics and Communication Engineering",
    "Bachelor of Science in Industrial Automation and Mechatronics",
    "Bachelor of Science in Ceramics Engineering",
    "Bachelor of Science in Mechanical Engineering",
    "Bachelor of Engineering Technology Major in Chemical Engineering and Technology",
    "Bachelor of Engineering Technology Major in Civil Engineering Technology",
    "Bachelor of Engineering Technology Major in Electrical Engineering Technology",
    "Bachelor of Engineering Technology Major in Electronics Engineering Technology",
    "Bachelor of Engineering Technology Major in Metallurgical and Materials Engineering Technology",
    "Bachelor of Engineering Technology Major in Mechanical Engineering Technology",
    "Bachelor of Science in Biology",
    "Bachelor of Science in Chemistry",
    "Bachelor of Science in Mathematics",
    "Bachelor of Science in Statistics",
    "Bachelor of Science in Physics",
    "Bachelor of Science in Computer Science",
    "Bachelor of Science in Information Technology",
    "Bachelor of Science in Information System",
    "Bachelor of Science in Computer Application",
    "Bachelor of Elementary Education in Science and Mathematics",
    "Bachelor of Elementary Education in Language Education",
    "Bachelor of Secondary Education in Chemistry",
    "Bachelor of Secondary Education in Physics",
    "Bachelor of Secondary Education in Mathematics",
    "Bachelor of Secondary Education in Biology",
    "Bachelor of Secondary Education in Filipino",
    "Bachelor of Technology and Livelihood in Home Economics",
    "Bachelor of Technology and Livelihood in Industrial Arts",
    "Bachelor of Technical-Vocational Teacher Education in Drafting Technology",
    "Bachelor of Arts in English Language Studies",
    "Bachelor of Arts in Language and Culture Studies",
    "Bachelor of Arts in Filipino",
    "Bachelor of Arts in History",
    "Bachelor of Arts in Panitikan",
    "Bachelor of Arts in Political Science",
    "Bachelor of Arts in Sociology",
    "Bachelor of Arts in Psychology",
    "Bachelor of Science in Psychology",
    "Bachelor of Science in Philosophy",
    "Bachelor of Science in Accountancy",
    "Bachelor of Science in Economics",
    "Bachelor of Science in Business Administration Major in Business Economics",
    "Bachelor of Science in Entrepreneurship",
    "Bachelor of Science in Hotel Management",
    "Bachelor of Science in Nursing"]
        self.ui.GENDER_Box.addItems([" ", "Male", "Female", "Non-Binary"])
        self.ui.YEARLEVEL_Box.addItems([" ", "1st", "2nd", "3rd", "4th", "5th", "6th"])
        self.ui.COURSE_Box.addItems([" "] + courses)
        self.ui.GENDER2_Box.addItems([" ", "Male", "Female", "Non-Binary"])
        self.ui.YEARLEVEL2_Box.addItems([" ", "1st", "2nd", "3rd", "4th", "5th", "6th"])
        self.ui.COURSE2_Box.addItems([" "] + courses) 
                                     
                                     

if __name__ == "__main__":
    app = QApplication(sys.argv)

    with open("style.qss") as f:
        style_str = f.read()

  

    app.setStyleSheet(style_str)

    window = SSIS_Window()
    window.show()
    sys.exit(app.exec())